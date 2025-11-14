import psutil as util
import sched
import time
from openpyxl import Workbook
import matplotlib.pyplot as plt
import pandas as pd
import argparse

row = 2  # next Excel row to write (row 1 is header)

def sys_output(scheduler, start_time, duration):
    global row, sheet1, wb, excel_path

    current_time = time.time()

    # stop after 10 seconds
    if current_time - start_time >= duration:
        wb.save(excel_path)
        return

    # capture metrics
    cpu_percent = util.cpu_percent(interval=1)
    memory_percent = util.virtual_memory().percent
    disk_percent = util.disk_usage('C:\\').percent

    # write one sample per row: (Time, CPU, Memory, Disk)
    sheet1.cell(row=row, column=1, value=round(current_time - start_time))
    sheet1.cell(row=row, column=2, value=cpu_percent)
    sheet1.cell(row=row, column=3, value=memory_percent)
    sheet1.cell(row=row, column=4, value=disk_percent)

    row += 1

    # save workbook frequently
    wb.save(excel_path)

    # schedule next run in 5 seconds
    scheduler.enter(5, 1, sys_output, (scheduler, start_time, duration))

def cpu_percent_graph(df):
    plt.figure(figsize=(8, 5))

    x = df["Time (sec)"]
    y = df["CPU Percentage"]

    plt.scatter(x, y)
    plt.plot(x,y)
    plt.xlabel("Time (sec)")
    plt.ylabel("CPU Percentage")
    plt.title("CPU Percentage vs. Time")
    plt.show()

def memory_percent_graph(df):
    plt.figure(figsize=(8, 5))

    x = df["Time (sec)"]
    y = df["Memory Percentage"]

    plt.scatter(x, y)
    plt.plot(x,y)
    plt.xlabel("Time (sec)")
    plt.ylabel("Memory Percentage")
    plt.title("Memory Percentage vs. Time")
    plt.show()

def disk_percent_graph(df):
    plt.figure(figsize=(8, 5))

    x = df["Time (sec)"]
    y = df["Disk Percentage"]

    plt.scatter(x, y)
    plt.plot(x,y)
    plt.xlabel("Time (sec)")
    plt.ylabel("Disk Percentage")
    plt.title("Disk Percentage vs. Time")
    plt.show()

if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='System Resource Monitor')

    parser.add_argument('--duration', type=int)
    parser.add_argument('--location', type=str)

    args = parser.parse_args()

    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Sheet 1"

    # header row
    sheet1.cell(row=1, column=1, value='Time (sec)')
    sheet1.cell(row=1, column=2, value='CPU Percentage')
    sheet1.cell(row=1, column=3, value='Memory Percentage')
    sheet1.cell(row=1, column=4, value='Disk Percentage')

    excel_path = args.location

    s = sched.scheduler(time.time, time.sleep)

    start_time = time.time()
    s.enter(0, 1, sys_output, (s, start_time, args.duration))
    s.run()

    wb.save(excel_path)

    df = pd.read_excel(excel_path)

    cpu_percent_graph(df)
    memory_percent_graph(df)
    disk_percent_graph(df)
