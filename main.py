import sched
import time
from openpyxl import Workbook
import pandas as pd
import argparse
import monitor
from flask import Flask, jsonify
import plotting
import threading

app = Flask(__name__)

latest_metrics = {
    "time": 0,
    "CPU Percent": 0,
    "Memory Percent": 0,
    "Disk Percent": 0
}

@app.route('/monitor')
def penis():
    return jsonify(latest_metrics)

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="System Resource Monitor")
    parser.add_argument("--duration", type=int)
    parser.add_argument("--excel_location", type=str)
    parser.add_argument("--json_location", type=str)
    parser.add_argument("--database_location", type=str)
    args = parser.parse_args()

    # -------------------------------
    # EXCEL SETUP  (RESTORED)
    # -------------------------------
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Sheet 1"

    sheet1.cell(row=1, column=1, value="Time (sec)")
    sheet1.cell(row=1, column=2, value="CPU Percentage")
    sheet1.cell(row=1, column=3, value="Memory Percentage")
    sheet1.cell(row=1, column=4, value="Disk Percentage")

    excel_path = args.excel_location
    json_path = args.json_location
    db_path = args.database_location

    open(json_path, 'w').close()

    s = sched.scheduler(time.time, time.sleep)
    start_time = time.time()

    # --------------------------------------------
    # RUN sys_output WITH EXCEL SUPPORT RESTORED
    # --------------------------------------------
    s.enter(
        0,
        1,
        monitor.sys_output,
        (s, start_time, args.duration, json_path, wb, sheet1, excel_path)
    )

    # -------------------------------
    # THREAD 1: LIVE MONITOR LOOP
    # -------------------------------
    t1 = threading.Thread(
        target=monitor.monitor_loop,
        args=(latest_metrics,),
        daemon=True
    )
    t1.start()

    # -------------------------------
    # THREAD 2: FLASK SERVER
    # -------------------------------
    t2 = threading.Thread(
        target=app.run,
        kwargs={"debug": True, "use_reloader": False},
        daemon=True
    )
    t2.start()

    # -------------------------------
    # RUN PHASE 1 COLLECTION
    # -------------------------------
    s.run()

    # -------------------------------
    # SQLITE LOGGING
    # -------------------------------
    monitor.write_to_db(json_path, db_path)

    # -------------------------------
    # SAVE EXCEL FILE
    # -------------------------------
    wb.save(excel_path)

    # -------------------------------
    # GRAPHING
    # -------------------------------
    df = pd.read_excel(excel_path)
    plotting.cpu_percent_graph(df)
    plotting.memory_percent_graph(df)
    plotting.disk_percent_graph(df)
